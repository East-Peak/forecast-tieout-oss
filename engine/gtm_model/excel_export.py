"""
Excel Export Module — Planning Tie-Out Workbook Generator.

Generates a multi-sheet Excel workbook from a TieoutResult with:
- Executive Summary (FY26 overview, key metrics, health indicator)
- Bookings Bridge (monthly stacked area — existing wins vs future generation)
- Funnel Tie-Out (stage-by-stage plan vs actual/projected comparison)
- Operating Reforecast (quarter-to-date actuals vs plan pace vs BU)
- Source Detail (monthly driver drilldown by modeled source stream)
- Expansion Detail (quarter-level existing-customer growth workstream)
- Pipeline Cohorts (monthly pipeline creation → expected bookings)
- Monthly Capacity (headcount, ramp, capacity timeline + SE section)
- Scenario Planner (optional — per-quarter overrides and comparison)
- Scenarios (if multiple scenarios computed)
- Data Health (freshness, bookings reconciliation, decay curve, targets)
- Assumptions (all model parameters for auditability)

Usage:
    from gtm_model.excel_export import export_tieout_workbook
    from gtm_model.tieout import PlanningTieout

    tieout = PlanningTieout()
    result = tieout.compute_full()
    path = export_tieout_workbook(result, "planning_tieout.xlsx")
"""

import io
import json
import subprocess
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Union

from gtm_model.tieout.views.recommendations import format_money
from gtm_model.tieout.types import ScenarioResult, TieoutResult
from gtm_model.tieout.views.view_models import (
    build_bookings_bridge_view_model,
    build_funnel_pacing_view_model,
    build_scenario_overlay_view_model,
    build_scenario_override_rows,
    build_se_capacity_view_model,
)

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import kaleido  # noqa: F401 — presence check for Plotly image export
    HAS_KALEIDO = True
except ImportError:
    HAS_KALEIDO = False


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

CELL_BG = "FFFFFF"
HEADER_BG = "4472C4"
ACCENT_BLUE = "2F5496"
GREEN = "548235"
YELLOW = "BF8F00"
RED = "C00000"
CELL_FONT_COLOR = "333333"
HEADER_FONT_COLOR = "FFFFFF"
BORDER_COLOR = "D6DCE4"

# Status → color mapping
_STATUS_COLOR = {
    "green": GREEN,
    "yellow": YELLOW,
    "red": RED,
    "ok": GREEN,
    "info": YELLOW,
    "warning": RED,
    "aligned": GREEN,
    "diverged": RED,
    "committed": GREEN,
    "building": YELLOW,
    "planned": "2F5496",
}

_SOURCE_DISPLAY_NAMES = {
    "marketing_sdr": "Marketing / SDR",
    "marketing / sdr": "Marketing / SDR",
    "marketing sdr": "Marketing / SDR",
    "ae_selfgen": "AE Self-Gen",
    "ae selfgen": "AE Self-Gen",
    "ae self gen": "AE Self-Gen",
    "ae self-gen": "AE Self-Gen",
    "plg": "PLG",
}

_PROVENANCE_LABELS = {
    "observed": "Observed",
    "blended": "Blended",
    "plan": "Plan",
}

_MODE_LABELS = {
    "cdw": "warehouse",
    "config": "Config",
}

_CONFIDENCE_LABELS = {
    "committed": "Committed",
    "building": "Building",
    "planned": "Planned",
}

_ACRONYM_LABELS = {
    "cdw": "warehouse",
    "plg": "PLG",
    "arr": "ARR",
    "nacv": "NACV",
    "mql": "MQL",
    "sdr": "SDR",
    "ae": "AE",
    "qtd": "QTD",
    "fy": "FY",
    "sfdc": "SFDC",
    "api": "API",
}

_TRANSITION_LABELS = {
    "mql_to_s0": "MQL -> S0",
    "s0_to_s1": "S0 -> S1",
    "s1_to_s2": "S1 -> S2",
}


def _primary_scenario(result: TieoutResult) -> ScenarioResult:
    """Return the default user-facing scenario for exports."""
    return getattr(result, "primary_scenario", None) or getattr(result, "trajectory", None) or result.base


def _archived_plan(result: TieoutResult) -> ScenarioResult:
    """Return the archived plan scenario for compatibility exports."""
    return getattr(result, "archived_plan", None) or result.base


def _display_label(value: object, *, title_case: bool = True) -> str:
    """Render compact display labels without mangling filenames."""
    raw_text = str(value or "").strip()
    text = raw_text.replace("_", " ")
    if not text:
        return ""
    if raw_text.lower() in _ACRONYM_LABELS:
        return _ACRONYM_LABELS[raw_text.lower()]
    if any(token in raw_text for token in (".yaml", ".yml", ".xlsx", "/", "\\")):
        return raw_text
    return text.title() if title_case else text


def _display_source_name(value: object) -> str:
    """Normalize reviewer-facing source stream labels."""
    raw_text = str(value or "").strip()
    normalized = (
        raw_text.lower()
        .replace("-", " ")
        .replace("_", " ")
        .replace("  ", " ")
        .strip()
    )
    normalized = " / ".join(part.strip() for part in normalized.split("/")) if "/" in normalized else normalized
    return _SOURCE_DISPLAY_NAMES.get(normalized, _display_label(raw_text))


def _display_provenance(value: object) -> str:
    """Normalize provenance labels."""
    raw_text = str(value or "").strip().lower()
    return _PROVENANCE_LABELS.get(raw_text, _display_label(value))


def _display_mode(value: object) -> str:
    """Normalize source-mode labels."""
    raw_text = str(value or "").strip().lower()
    return _MODE_LABELS.get(raw_text, _display_label(value))


def _display_confidence_tier(value: object) -> str:
    """Normalize confidence-tier labels."""
    raw_text = str(value or "").strip().lower()
    return _CONFIDENCE_LABELS.get(raw_text, _display_label(value))


def _display_transition_label(value: object) -> str:
    """Normalize funnel transition labels."""
    raw_text = str(value or "").strip().lower()
    return _TRANSITION_LABELS.get(raw_text, _display_label(value))


def _render_chart_to_image(fig, width=800, height=400):
    """Render a Plotly figure to PNG bytes for Excel embedding.

    Returns a BytesIO stream, or None if kaleido is not installed.
    """
    if not HAS_KALEIDO:
        return None
    try:
        img_bytes = fig.to_image(format="png", width=width, height=height, scale=2)
        return io.BytesIO(img_bytes)
    except Exception:
        return None


def _build_bookings_bridge_chart(scenario) -> "object | None":
    """Build a Plotly stacked area chart for the Bookings Bridge sheet."""
    if not HAS_KALEIDO:
        return None
    try:
        import plotly.graph_objects as go
    except ImportError:
        return None

    months = scenario.monthly_months or []
    existing_wins = scenario.monthly_existing_inventory_wins or []
    future_wins = scenario.monthly_future_generation_wins or []
    monthly_cap = scenario.monthly_capacity or []

    if not months:
        return None

    labels = [m.strftime("%b %Y") if hasattr(m, "strftime") else str(m) for m in months]
    targets = [
        monthly_cap[i].monthly_target if i < len(monthly_cap) and monthly_cap[i].monthly_target is not None else None
        for i in range(len(months))
    ]
    ex = [existing_wins[i] if i < len(existing_wins) else 0 for i in range(len(months))]
    fu = [future_wins[i] if i < len(future_wins) else 0 for i in range(len(months))]

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=labels, y=ex, name="Existing Inventory",
        fill="tozeroy", mode="lines",
        line=dict(color="#4472C4"),
        fillcolor="rgba(68,114,196,0.5)",
    ))
    fig.add_trace(go.Scatter(
        x=labels, y=[e + f for e, f in zip(ex, fu)], name="Future Generation",
        fill="tonexty", mode="lines",
        line=dict(color="#548235"),
        fillcolor="rgba(84,130,53,0.5)",
    ))
    fig.add_trace(go.Scatter(
        x=labels, y=targets, name="Target",
        mode="lines+markers",
        line=dict(color="#C00000", dash="dash", width=2),
    ))
    fig.update_layout(
        title="Monthly Bookings Bridge — Existing vs Future vs Target",
        xaxis_title="Month", yaxis_title="ARR ($)",
        yaxis_tickprefix="$", yaxis_tickformat=",",
        template="plotly_white",
        plot_bgcolor="white", paper_bgcolor="white",
        width=800, height=400,
        margin=dict(l=60, r=30, t=50, b=50),
    )
    return fig


def _build_bookings_cumulative_chart(scenario) -> "object | None":
    """Build a Plotly cumulative bookings stacked area chart."""
    if not HAS_KALEIDO:
        return None
    try:
        import plotly.graph_objects as go
    except ImportError:
        return None

    months = scenario.monthly_months or []
    existing_wins = scenario.monthly_existing_inventory_wins or []
    future_wins = scenario.monthly_future_generation_wins or []
    monthly_cap = scenario.monthly_capacity or []

    if not months:
        return None

    n = len(months)
    labels = [m.strftime("%b %Y") if hasattr(m, "strftime") else str(m) for m in months]
    targets = [
        monthly_cap[i].monthly_target if i < len(monthly_cap) and monthly_cap[i].monthly_target is not None else None
        for i in range(n)
    ]

    # Cumulative sums
    cum_existing = []
    cum_total = []
    cum_target = []
    target_supported = any(target is not None for target in targets)
    run_e, run_t, run_tgt = 0, 0, 0
    for i in range(n):
        e = existing_wins[i] if i < len(existing_wins) else 0
        f = future_wins[i] if i < len(future_wins) else 0
        run_e += e
        run_t += e + f
        cum_existing.append(run_e)
        cum_total.append(run_t)
        if target_supported:
            run_tgt += targets[i] or 0
            cum_target.append(run_tgt)
        else:
            cum_target.append(None)

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=labels, y=cum_existing, name="Existing Pipeline (Cumulative)",
        fill="tozeroy",
        line=dict(color="rgba(88, 166, 255, 0.9)", width=1),
        fillcolor="rgba(88, 166, 255, 0.35)",
    ))
    fig.add_trace(go.Scatter(
        x=labels, y=cum_total, name="+ Future Pipeline (Cumulative)",
        fill="tonexty",
        line=dict(color="rgba(63, 185, 80, 0.9)", width=1),
        fillcolor="rgba(63, 185, 80, 0.35)",
    ))
    fig.add_trace(go.Scatter(
        x=labels, y=cum_target, name="Cumulative Target",
        line=dict(color="#C00000", dash="dash", width=2.5),
    ))
    fig.update_layout(
        title="Cumulative Bookings Path",
        xaxis_title="Month", yaxis_title="Cumulative Sales-Led Bookings",
        yaxis_tickprefix="$", yaxis_tickformat=",",
        template="plotly_white",
        plot_bgcolor="white", paper_bgcolor="white",
        width=800, height=400,
        margin=dict(l=60, r=30, t=50, b=50),
    )
    return fig


def _build_scenario_overlay_chart(baseline, flexed) -> "object | None":
    """Build a Plotly grouped bar chart comparing baseline vs scenario by quarter."""
    if not HAS_KALEIDO or flexed is None:
        return None
    try:
        import plotly.graph_objects as go
    except ImportError:
        return None

    quarters = [q.quarter for q in flexed.quarters]
    baseline_vals = [
        baseline.quarters[i].bu_sales_led_arr if i < len(baseline.quarters) else 0
        for i in range(len(quarters))
    ]
    scenario_vals = [q.bu_sales_led_arr for q in flexed.quarters]
    target_vals = [q.td_bookings for q in flexed.quarters]

    if not quarters:
        return None

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=quarters, y=target_vals, name="Target",
        marker_color="#D6DCE4",
    ))
    fig.add_trace(go.Bar(
        x=quarters, y=baseline_vals, name="Baseline",
        marker_color="#4472C4",
    ))
    fig.add_trace(go.Bar(
        x=quarters, y=scenario_vals, name="Scenario",
        marker_color="#548235",
    ))
    fig.update_layout(
        title="Scenario Comparison — Sales-Led ARR by Quarter",
        xaxis_title="Quarter", yaxis_title="Sales-Led ARR ($)",
        yaxis_tickprefix="$", yaxis_tickformat=",",
        barmode="group",
        template="plotly_white",
        plot_bgcolor="white", paper_bgcolor="white",
        width=800, height=400,
        margin=dict(l=60, r=30, t=50, b=50),
    )
    return fig


def _setup_styles(wb: "Workbook"):
    """Register reusable styles on the workbook."""
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # Header style
    header = NamedStyle(name="header_style")
    header.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header.border = thin_border
    wb.add_named_style(header)

    # Currency style
    currency = NamedStyle(name="currency_style")
    currency.number_format = '$#,##0'
    currency.font = Font(color=CELL_FONT_COLOR)
    currency.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    currency.border = thin_border
    wb.add_named_style(currency)

    # Percentage style
    pct = NamedStyle(name="pct_style")
    pct.number_format = '0.0%'
    pct.font = Font(color=CELL_FONT_COLOR)
    pct.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    pct.border = thin_border
    wb.add_named_style(pct)

    # Number style
    num = NamedStyle(name="num_style")
    num.number_format = '#,##0'
    num.font = Font(color=CELL_FONT_COLOR)
    num.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    num.border = thin_border
    wb.add_named_style(num)

    # Decimal number style
    decimal = NamedStyle(name="decimal_style")
    decimal.number_format = '#,##0.0'
    decimal.font = Font(color=CELL_FONT_COLOR)
    decimal.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    decimal.border = thin_border
    wb.add_named_style(decimal)

    # Label style
    label = NamedStyle(name="label_style")
    label.font = Font(color=CELL_FONT_COLOR, size=11)
    label.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    label.border = thin_border
    label.alignment = Alignment(horizontal="left")
    wb.add_named_style(label)

    # Title style
    title = NamedStyle(name="title_style")
    title.font = Font(bold=True, color=ACCENT_BLUE, size=14)
    title.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    wb.add_named_style(title)

    # Subtitle
    subtitle = NamedStyle(name="subtitle_style")
    subtitle.font = Font(bold=True, color=CELL_FONT_COLOR, size=12)
    subtitle.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    wb.add_named_style(subtitle)

    # Green (positive)
    green = NamedStyle(name="green_currency")
    green.number_format = '$#,##0'
    green.font = Font(color=GREEN)
    green.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    green.border = thin_border
    wb.add_named_style(green)

    # Red (negative/gap)
    red = NamedStyle(name="red_currency")
    red.number_format = '$#,##0'
    red.font = Font(color=RED)
    red.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    red.border = thin_border
    wb.add_named_style(red)


def _apply_cell_bg(ws, max_row: int, max_col: int):
    """Apply white background to all unstyled cells in range."""
    fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.fill.start_color.rgb == "00000000":
                cell.fill = fill


def _auto_width(ws, min_width: int = 12, max_width: int = 25):
    """Auto-size column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


def _status_label_cell(ws, row: int, col: int, status: str):
    """Write a colored status badge ('● green', etc.) into a cell."""
    color = _STATUS_COLOR.get(status, CELL_FONT_COLOR)
    cell = ws.cell(row=row, column=col, value=f"● {status.upper()}")
    cell.font = Font(bold=True, color=color)
    cell.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
    return cell


def _get_git_sha() -> str:
    """Return the short git SHA if available."""
    repo_root = Path(__file__).resolve().parent.parent
    try:
        result = subprocess.run(
            ["git", "-C", str(repo_root), "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            timeout=5,
            check=True,
        )
        return result.stdout.strip() or "unknown"
    except Exception:
        return "unknown"


def _comparison_contract() -> dict:
    """Return the exported plan-comparison contract used across Python outputs."""
    return {
        "operator_comparable_metric": "sales_led_arr",
        "operator_comparable_label": "Sales-Led ARR",
        "operator_comparable_role": "primary_comparison",
        "executive_context_metric": "total_net_new_arr",
        "executive_context_label": "Total Net New ARR",
        "executive_context_role": "secondary_reference_only",
        "note": (
            "Sales-led ARR is the operator-comparable comparison rail. "
            "Total net new ARR remains executive context only."
        ),
    }


def _build_review_metadata(result: TieoutResult) -> dict:
    """Build metadata attached to review-pack exports."""
    scenario = _primary_scenario(result)
    modes = sorted({
        str(row.get("mode", "unknown"))
        for row in (getattr(scenario, "monthly_source_detail", []) or [])
    })
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "git_sha": _get_git_sha(),
        "overall_health": (result.health_status or {}).get("overall_status"),
        "top_down_plan": result.top_down_plan or {},
        "monthly_source_modes": modes,
        "comparison_contract": _comparison_contract(),
    }


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_executive_summary(wb: "Workbook", result: TieoutResult):
    """Build the Executive Summary sheet with health indicator and funnel gap summary."""
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = ACCENT_BLUE

    base = _archived_plan(result)
    scenario = _primary_scenario(result)
    health = result.health_status or {}
    plan = result.top_down_plan or {}
    comparison_contract = _comparison_contract()

    # Title
    ws["A1"] = "Forecast Tieout — Executive Summary"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    row = 4
    if plan:
        ws.cell(row=row, column=1, value="TOP-DOWN PLAN").style = "subtitle_style"
        ws.merge_cells(f"A{row}:G{row}")
        row += 1

        plan_rows = [
            ("Plan Label", plan.get("label", "")),
            ("Plan Role", str(plan.get("status", "")).replace("_", " ").title()),
            ("Derived From", plan.get("derived_from", "")),
            ("Source Workbook", plan.get("source_workbook", "")),
            ("Source Context", plan.get("source_context", "")),
            ("Guidance", plan.get("guidance", "")),
            ("Comparison Contract", comparison_contract["note"]),
        ]
        for label, value in plan_rows:
            if not value:
                continue
            ws.cell(row=row, column=1, value=label).style = "label_style"
            ws.cell(row=row, column=2, value=value).style = "label_style"
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            row += 1

    # ── Data Health Status Banner ──────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="DATA HEALTH").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    overall_status = health.get("overall_status", "unknown")
    ws.cell(row=row, column=1, value="Overall Health Status").style = "label_style"
    _status_label_cell(ws, row, 2, overall_status)
    row += 1

    # Freshness sub-status
    freshness = health.get("freshness", {})
    if freshness:
        ws.cell(row=row, column=1, value="  Data Freshness").style = "label_style"
        _status_label_cell(ws, row, 2, freshness.get("status", "unknown"))
        msg = freshness.get("message", "")
        if msg:
            ws.cell(row=row, column=3, value=msg[:80]).style = "label_style"
        row += 1

    # ── Key Metrics ────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="KEY METRICS").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    fy_sales_led_deficit = base.fy_bookings_td - scenario.fy_bookings_bu
    fy_sales_led_deficit_pct = fy_sales_led_deficit / base.fy_bookings_td if base.fy_bookings_td else 0
    metrics = [
        ("Beginning ARR (Live)", result.beginning_arr, "currency_style"),
        ("Board-Plan Beginning ARR", result.top_down_beginning_arr, "currency_style"),
        ("Ending ARR Plan", scenario.fy_ending_arr_td, "currency_style"),
        ("Plan Sales-Led ARR (Comparable)", base.fy_bookings_td, "currency_style"),
        ("Trajectory Sales-Led ARR", scenario.fy_bookings_bu, "currency_style"),
        (
            "FY Sales-Led ARR Deficit vs Plan",
            fy_sales_led_deficit,
            "red_currency" if fy_sales_led_deficit > 0 else "green_currency",
        ),
        ("FY Sales-Led Deficit %", fy_sales_led_deficit_pct, "pct_style"),
        ("Plan Total Net New ARR (Executive Context)", base.fy_total_td, "currency_style"),
        ("Trajectory Total Net New ARR (Executive Context)", scenario.fy_total_bu, "currency_style"),
        (
            "FY Executive-Context Deficit vs Plan",
            scenario.fy_gap,
            "red_currency" if scenario.fy_gap > 0 else "green_currency",
        ),
        ("FY Executive-Context Deficit %", scenario.fy_gap_pct, "pct_style"),
    ]

    for label, value, style in metrics:
        ws.cell(row=row, column=1, value=label).style = "label_style"
        cell = ws.cell(row=row, column=2, value=value)
        cell.style = style
        row += 1

    beginning_arr_meta = getattr(result, "beginning_arr_provenance", {}) or {}
    if beginning_arr_meta:
        ws.cell(row=row, column=1, value="Beginning ARR Source").style = "label_style"
        ws.cell(row=row, column=2, value=_display_label(beginning_arr_meta.get("source", "unknown"))).style = "label_style"
        row += 1
        ws.cell(row=row, column=1, value="Beginning ARR Method").style = "label_style"
        ws.cell(row=row, column=2, value=_display_label(beginning_arr_meta.get("method", "unknown"))).style = "label_style"
        row += 1
        if beginning_arr_meta.get("warning"):
            ws.cell(row=row, column=1, value="Beginning ARR Note").style = "label_style"
            ws.cell(row=row, column=2, value=beginning_arr_meta.get("warning")).style = "label_style"
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            row += 1

    bookings_summary = getattr(result, "bookings_summary", {}) or {}
    bookings_totals = bookings_summary.get("totals", {}) if isinstance(bookings_summary, dict) else {}
    if bookings_totals:
        ws.cell(row=row, column=1, value="FY26 YTD Closed-Won Gross Bookings").style = "label_style"
        ws.cell(row=row, column=2, value=float(bookings_totals.get("amount") or 0.0)).style = "currency_style"
        row += 1
        ws.cell(row=row, column=1, value="FY26 YTD Closed-Won Recurring Bookings (Y1 ARR)").style = "label_style"
        ws.cell(row=row, column=2, value=float(bookings_totals.get("year1_arr") or 0.0)).style = "currency_style"
        row += 1
        ws.cell(row=row, column=1, value="FY26 YTD Closed-Won NACV").style = "label_style"
        ws.cell(row=row, column=2, value=float(bookings_totals.get("nacv") or 0.0)).style = "currency_style"
        row += 1
        ws.cell(row=row, column=1, value="FY26 YTD Closed-Won Non-Recurring Bookings").style = "label_style"
        ws.cell(row=row, column=2, value=float(bookings_totals.get("non_recurring") or 0.0)).style = "currency_style"
        row += 1
        ws.cell(row=row, column=1, value="Closed-Won Opportunity Count").style = "label_style"
        ws.cell(row=row, column=2, value=int(bookings_totals.get("won_count") or 0)).style = "num_style"
        row += 1

    bookings_meta = getattr(result, "bookings_summary_provenance", {}) or {}
    if bookings_meta:
        ws.cell(row=row, column=1, value="Bookings Source").style = "label_style"
        ws.cell(row=row, column=2, value=_display_label(bookings_meta.get("source", "unknown"))).style = "label_style"
        row += 1
        ws.cell(row=row, column=1, value="Bookings Method").style = "label_style"
        ws.cell(row=row, column=2, value=_display_label(bookings_meta.get("method", "unknown"))).style = "label_style"
        row += 1
        if bookings_meta.get("warning"):
            ws.cell(row=row, column=1, value="Bookings Note").style = "label_style"
            ws.cell(row=row, column=2, value=bookings_meta.get("warning")).style = "label_style"
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            row += 1

    # ── Funnel Gap Summary ─────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="FUNNEL GAP SUMMARY (Q1FY26)").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    funnel_headers = ["Stage", "Plan", "Actual / Trajectory", "Delta", "Delta %"]
    for col, h in enumerate(funnel_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    # Pull funnel tieout from first (most recent/committed) quarter
    q1 = scenario.quarters[0] if scenario.quarters else None
    if q1:
        ft = q1.funnel_tieout or {}
        funnel_stages = [
            ("MQLs/week",    ft.get("mqls_weekly", {}), False),
            ("S0/week",      ft.get("s0_weekly", {}),   False),
            ("S1/week",      ft.get("s1_weekly", {}),   False),
            ("S2 created/week", ft.get("s2_weekly", {}), False),
        ]
        for stage_label, stage_data, is_currency in funnel_stages:
            if not stage_data:
                continue
            row += 1
            plan_val = stage_data.get("plan", 0)
            actual_val = stage_data.get("actual", 0)
            delta_val = stage_data.get("delta", actual_val - plan_val)
            delta_pct = delta_val / plan_val if plan_val else 0

            ws.cell(row=row, column=1, value=stage_label).style = "label_style"
            ws.cell(row=row, column=2, value=plan_val).style = "num_style"
            ws.cell(row=row, column=3, value=actual_val).style = "num_style"
            gap_style = "green_currency" if delta_val >= 0 else "red_currency"
            delta_cell = ws.cell(row=row, column=4, value=delta_val)
            delta_cell.style = gap_style
            delta_cell.number_format = '+#,##0;-#,##0;0'
            ws.cell(row=row, column=5, value=delta_pct).style = "pct_style"

    # ── Quarterly Breakdown ────────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="QUARTERLY ARR TIE-OUT").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    headers = [
        "Quarter",
        "Plan Sales-Led ARR (Comparable)",
        "Trajectory Sales-Led ARR",
        "Sales-Led ARR Deficit vs Plan",
        "Sales-Led Deficit %",
        "Plan Total Net New ARR (Executive Context)",
        "Trajectory Total Net New ARR (Executive Context)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for q in scenario.quarters:
        row += 1
        ws.cell(row=row, column=1, value=q.quarter).style = "label_style"
        ws.cell(row=row, column=2, value=q.td_bookings).style = "currency_style"
        ws.cell(row=row, column=3, value=q.bu_sales_led_arr).style = "currency_style"
        gap_style = "red_currency" if q.bookings_gap > 0 else "green_currency"
        ws.cell(row=row, column=4, value=q.bookings_gap).style = gap_style
        ws.cell(row=row, column=5, value=q.bookings_gap_pct).style = "pct_style"
        ws.cell(row=row, column=6, value=q.td_total_net_new).style = "currency_style"
        ws.cell(row=row, column=7, value=q.bu_total_arr).style = "currency_style"

    # FY Total row
    row += 1
    ws.cell(row=row, column=1, value="FY26 TOTAL").style = "header_style"
    ws.cell(row=row, column=2, value=base.fy_bookings_td).style = "currency_style"
    ws.cell(row=row, column=3, value=scenario.fy_bookings_bu).style = "currency_style"
    fy_gap_style = "red_currency" if fy_sales_led_deficit > 0 else "green_currency"
    ws.cell(row=row, column=4, value=fy_sales_led_deficit).style = fy_gap_style
    fy_bookings_gap_pct = fy_sales_led_deficit_pct
    ws.cell(row=row, column=5, value=fy_bookings_gap_pct).style = "pct_style"
    ws.cell(row=row, column=6, value=base.fy_total_td).style = "currency_style"
    ws.cell(row=row, column=7, value=scenario.fy_total_bu).style = "currency_style"

    # ── Headcount Plan ─────────────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="HEADCOUNT PLAN").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    hc_headers = ["Quarter", "AEs (Target)", "AEs (BU)", "Ramped AEs",
                   "SEs", "SDRs", "Total GTM"]
    for col, h in enumerate(hc_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for q in scenario.quarters:
        row += 1
        ws.cell(row=row, column=1, value=q.quarter).style = "label_style"
        ws.cell(row=row, column=2, value=q.td_aes).style = "num_style"
        ws.cell(row=row, column=3, value=q.bu_total_aes).style = "num_style"
        ws.cell(row=row, column=4, value=q.bu_ramped_aes).style = "num_style"
        ws.cell(row=row, column=5, value=q.td_ses).style = "num_style"
        ws.cell(row=row, column=6, value=q.td_sdrs).style = "num_style"
        ws.cell(row=row, column=7, value=q.td_total_gtm).style = "num_style"

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 7)


def _build_bookings_bridge(wb: "Workbook", result: TieoutResult):
    """Build the Bookings Bridge sheet — monthly existing vs future wins stacked breakdown."""
    ws = wb.create_sheet("Bookings Bridge")
    ws.sheet_properties.tabColor = "58A6FF"  # accent blue — hero sheet

    view_model = build_bookings_bridge_view_model(result)
    scenario = view_model["scenario"]

    ws["A1"] = "FY26 Bookings Bridge — Existing Inventory vs Future Generation"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    # ── Key Metrics Row ─────────────────────────────────────────────
    row = 4
    ws.cell(row=row, column=1, value="KEY METRICS").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    fy_target = view_model["totals"]["target"]
    fy_trajectory = view_model["totals"]["trajectory"]
    fy_gap = view_model["totals"]["gap"]
    fy_gap_pct = fy_gap / fy_target if fy_target else 0

    metrics = [
        ("FY Target (Sales-Led ARR)", fy_target, "currency_style"),
        ("Trajectory Forecast", fy_trajectory, "currency_style"),
        ("Gap", fy_gap, "red_currency" if fy_gap < 0 else "green_currency"),
        ("Gap %", fy_gap_pct, "pct_style"),
    ]
    for label, value, style in metrics:
        ws.cell(row=row, column=1, value=label).style = "label_style"
        ws.cell(row=row, column=2, value=value).style = style
        row += 1

    # ── Monthly Breakdown Table ─────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="MONTHLY BREAKDOWN").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    monthly_headers = ["Month", "Existing Wins", "Future Wins", "Total", "Target", "Gap"]
    for col, h in enumerate(monthly_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for row_data in view_model["monthly"]:
        row += 1
        gap = row_data["gap"]
        ws.cell(row=row, column=1, value=row_data["label"]).style = "label_style"
        ws.cell(row=row, column=2, value=row_data["existing_wins"]).style = "currency_style"
        ws.cell(row=row, column=3, value=row_data["future_wins"]).style = "currency_style"
        ws.cell(row=row, column=4, value=row_data["total_uncapped"]).style = "currency_style"
        ws.cell(row=row, column=5, value=row_data["target"] if row_data["target"] is not None else "—").style = (
            "currency_style" if row_data["target"] is not None else "label_style"
        )
        ws.cell(row=row, column=6, value=gap if gap is not None else "—").style = (
            "red_currency" if gap is not None and gap < 0 else "green_currency" if gap is not None else "label_style"
        )

    # FY Total row
    if view_model["monthly"]:
        row += 1
        ws.cell(row=row, column=1, value="FY26 TOTAL").style = "header_style"
        ws.cell(row=row, column=2, value=view_model["totals"]["existing"]).style = "currency_style"
        ws.cell(row=row, column=3, value=view_model["totals"]["future"]).style = "currency_style"
        ws.cell(row=row, column=4, value=view_model["totals"]["trajectory_uncapped"]).style = "currency_style"
        ws.cell(row=row, column=5, value=view_model["totals"]["target"]).style = "currency_style"

    # ── Quarterly Summary ───────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="QUARTERLY SUMMARY").style = "subtitle_style"
    ws.merge_cells(f"A{row}:H{row}")

    row += 1
    q_headers = ["Quarter", "Sales-Led Target", "Trajectory", "Gap", "Gap %",
                  "From Existing", "From Future", "Data Basis"]
    for col, h in enumerate(q_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for q_row in view_model["quarterly"]:
        row += 1
        q_gap = q_row["gap"]
        q_gap_pct = q_row["gap_pct"]
        data_basis = _display_confidence_tier(q_row["data_basis"])

        ws.cell(row=row, column=1, value=q_row["quarter"]).style = "label_style"
        ws.cell(row=row, column=2, value=q_row["target"]).style = "currency_style"
        ws.cell(row=row, column=3, value=q_row["trajectory"]).style = "currency_style"
        ws.cell(row=row, column=4, value=q_gap).style = "red_currency" if q_gap < 0 else "green_currency"
        ws.cell(row=row, column=5, value=q_gap_pct).style = "pct_style"
        ws.cell(row=row, column=6, value=q_row["from_existing"]).style = "currency_style"
        ws.cell(row=row, column=7, value=q_row["from_future"]).style = "currency_style"
        ws.cell(row=row, column=8, value=data_basis).style = "label_style"

    # ── Embedded Charts ──────────────────────────────────────────────
    if HAS_KALEIDO:
        row += 3
        ws.cell(row=row, column=1, value="CHARTS").style = "subtitle_style"
        ws.merge_cells(f"A{row}:H{row}")
        row += 1

        bridge_fig = _build_bookings_bridge_chart(scenario)
        img_stream = _render_chart_to_image(bridge_fig)
        if img_stream:
            img = XlImage(img_stream)
            img.width = 700
            img.height = 350
            ws.add_image(img, f"A{row}")
            row += 20

        cum_fig = _build_bookings_cumulative_chart(scenario)
        cum_stream = _render_chart_to_image(cum_fig)
        if cum_stream:
            img2 = XlImage(cum_stream)
            img2.width = 700
            img2.height = 350
            ws.add_image(img2, f"A{row}")
            row += 20

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 8)


def _build_funnel_tieout(wb: "Workbook", result: TieoutResult):
    """Build the Funnel Tie-Out sheet — plan vs actual/projected by stage."""
    ws = wb.create_sheet("Funnel Tie-Out")
    ws.sheet_properties.tabColor = "E3B341"  # warm amber

    scenario = _primary_scenario(result)

    ws["A1"] = "FY26 Funnel Tie-Out — Weekly Funnel Pace and Source Conversions"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    # ── Stage-by-stage comparison (one section per quarter) ────────────
    STAGE_ROWS = [
        ("MQLs / wk",       "mqls_weekly",  False),
        ("S0 / wk",         "s0_weekly",    False),
        ("S1 / wk",         "s1_weekly",    False),
        ("S2 / wk",         "s2_weekly",    False),
        ("Sales-Led ARR",   "_bookings",    True),   # synthetic: from quarter
        ("AEs in Seat",     "_ae_capacity", False),  # from monthly capacity
    ]

    row = 4

    for q in scenario.quarters:
        ws.cell(row=row, column=1, value=f"── {q.quarter} ──").style = "subtitle_style"
        ws.merge_cells(f"A{row}:G{row}")
        row += 1

        # Column headers
        headers = ["Stage", "Plan", "Actual / Trajectory", "Delta", "Delta %", "Quarter", "Confidence"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).style = "header_style"
        row += 1

        ft = q.funnel_tieout or {}

        for stage_label, key, is_currency in STAGE_ROWS:
            if key == "_bookings":
                plan_val = q.td_bookings
                actual_val = q.bu_sales_led_arr
            elif key == "_ae_capacity":
                plan_val = q.td_aes
                actual_val = q.bu_total_aes
            else:
                stage_data = ft.get(key, {})
                if not stage_data:
                    continue
                plan_val = stage_data.get("plan", 0)
                actual_val = stage_data.get("actual", 0)

            delta_val = actual_val - plan_val
            delta_pct = delta_val / plan_val if plan_val else 0

            ws.cell(row=row, column=1, value=stage_label).style = "label_style"

            if is_currency:
                ws.cell(row=row, column=2, value=plan_val).style = "currency_style"
                ws.cell(row=row, column=3, value=actual_val).style = "currency_style"
                gap_style = "green_currency" if delta_val >= 0 else "red_currency"
                delta_cell = ws.cell(row=row, column=4, value=delta_val)
                delta_cell.style = gap_style
            else:
                numeric_style = "num_style"
                ws.cell(row=row, column=2, value=plan_val).style = numeric_style
                ws.cell(row=row, column=3, value=actual_val).style = numeric_style
                gap_color = GREEN if delta_val >= 0 else RED
                delta_cell = ws.cell(row=row, column=4, value=delta_val)
                delta_cell.font = Font(color=gap_color)
                delta_cell.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
                delta_cell.number_format = '+#,##0;-#,##0;0'

            ws.cell(row=row, column=5, value=delta_pct).style = "pct_style"
            ws.cell(row=row, column=6, value=q.quarter).style = "label_style"
            ws.cell(row=row, column=7, value=_display_confidence_tier(q.confidence_tier)).style = "label_style"
            row += 1

        row += 1  # blank separator between quarters

    # ── Conversion Rates Table ─────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="SOURCE-LEVEL CONVERSION RATES").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    conv_headers = ["Quarter", "Transition", "Source Stream", "Rate", "Sample n", "Provenance", "Confidence"]
    for col, h in enumerate(conv_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"
    row += 1

    for q in scenario.quarters:
        cr = q.conversion_rates or {}
        for stage_key in ("mql_to_s0", "s0_to_s1", "s1_to_s2"):
            stage_data = cr.get(stage_key, {})
            if not stage_data:
                # Config-only mode has a flat structure
                if isinstance(cr.get(stage_key), dict):
                    # flat: {"rate": x, "n": y, "source": z}
                    pass
                continue
            for source, sdata in stage_data.items():
                if not isinstance(sdata, dict):
                    continue
                if stage_key == "mql_to_s0" and source != "marketing_sdr":
                    continue
                ws.cell(row=row, column=1, value=q.quarter).style = "label_style"
                ws.cell(row=row, column=2, value=_display_transition_label(stage_key)).style = "label_style"
                ws.cell(row=row, column=3, value=_display_source_name(source)).style = "label_style"
                ws.cell(row=row, column=4, value=sdata.get("rate", 0)).style = "pct_style"
                ws.cell(row=row, column=5, value=sdata.get("n", 0)).style = "num_style"
                ws.cell(row=row, column=6, value=_display_provenance(sdata.get("source", "plan"))).style = "label_style"
                ws.cell(row=row, column=7, value=_display_confidence_tier(q.confidence_tier)).style = "label_style"
                row += 1

    row += 2
    ws.cell(row=row, column=1, value="TARGET PROVENANCE BY QUARTER").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    provenance_headers = ["Quarter", "Label", "Method", "Source", "Reference Quarter", "Approved", "Notes"]
    for col, header in enumerate(provenance_headers, 1):
        ws.cell(row=row, column=col, value=header).style = "header_style"
    row += 1

    for q in scenario.quarters:
        provenance = getattr(q, "target_provenance", None) or {}
        ws.cell(row=row, column=1, value=q.quarter).style = "label_style"
        ws.cell(row=row, column=2, value=provenance.get("label", "")).style = "label_style"
        ws.cell(row=row, column=3, value=_display_label(provenance.get("method", ""))).style = "label_style"
        ws.cell(row=row, column=4, value=_display_label(provenance.get("source", ""))).style = "label_style"
        ws.cell(row=row, column=5, value=provenance.get("reference_quarter", "")).style = "label_style"
        ws.cell(row=row, column=6, value="Yes" if provenance.get("approved") else "No").style = "label_style"
        ws.cell(row=row, column=7, value=provenance.get("notes", "")).style = "label_style"
        row += 1

    # ── QTD Stage-Level Pacing ───────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="QTD STAGE-LEVEL FUNNEL PACING").style = "subtitle_style"
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    pacing_headers = ["Quarter", "Stage", "Weekly Target", "Weekly Actual", "QTD Target", "QTD Actual", "Pacing %"]
    for col, h in enumerate(pacing_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"
    row += 1

    for q in scenario.quarters:
        pacing_vm = build_funnel_pacing_view_model(q, as_of=date.today())
        if pacing_vm["quarter_state"] != "in_progress":
            continue
        for stage_row in pacing_vm["rows"]:
            ws.cell(row=row, column=1, value=q.quarter).style = "label_style"
            ws.cell(row=row, column=2, value=stage_row["stage"]).style = "label_style"
            ws.cell(row=row, column=3, value=stage_row["weekly_target"]).style = "num_style"
            ws.cell(row=row, column=4, value=stage_row["weekly_actual"]).style = "num_style"
            ws.cell(row=row, column=5, value=stage_row["qtd_target"]).style = "num_style"
            ws.cell(row=row, column=6, value=stage_row["qtd_actual"]).style = "num_style"
            ws.cell(row=row, column=7, value=stage_row["pacing_pct"]).style = "pct_style"
            row += 1

        row += 1  # blank separator between quarters

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 7)


def _build_operating_reforecast(wb: "Workbook", result: TieoutResult):
    """Build Operating Reforecast sheet — actuals vs plan pace vs BU."""
    ws = wb.create_sheet("Operating Reforecast")
    ws.sheet_properties.tabColor = GREEN

    scenario = _primary_scenario(result)
    as_of = date.today()

    ws["A1"] = "FY26 Operating Reforecast — Actuals vs Plan Pace vs Trajectory"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:J1")

    ws["A2"] = f"As of: {as_of.strftime('%Y-%m-%d')}"
    ws["A2"].style = "label_style"

    row = 4
    headers = [
        "Quarter",
        "Status",
        "Elapsed %",
        "Actual Sales-Led ARR QTD",
        "Plan Sales-Led ARR QTD",
        "Pace Gap",
        "Remaining to Plan",
        "Remaining to Trajectory",
        "Latest Trajectory Reforecast",
        "Reforecast Gap %",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).style = "header_style"
    row += 1

    for quarter in scenario.quarters:
        summary = quarter.reforecast_summary(as_of=as_of)
        ws.cell(row=row, column=1, value=quarter.quarter).style = "label_style"
        ws.cell(row=row, column=2, value=summary.get("quarter_state", "").replace("_", " ").title()).style = "label_style"
        ws.cell(row=row, column=3, value=summary.get("elapsed_fraction", 0.0)).style = "pct_style"
        ws.cell(row=row, column=4, value=summary.get("actual_bookings", 0.0)).style = "currency_style"
        ws.cell(row=row, column=5, value=summary.get("plan_to_date_bookings", 0.0)).style = "currency_style"
        pace_gap = summary.get("pace_gap", 0.0)
        ws.cell(row=row, column=6, value=pace_gap).style = "green_currency" if pace_gap >= 0 else "red_currency"
        ws.cell(row=row, column=7, value=summary.get("remaining_plan_bookings", 0.0)).style = "currency_style"
        ws.cell(row=row, column=8, value=summary.get("remaining_bu_bookings", 0.0)).style = "currency_style"
        ws.cell(row=row, column=9, value=summary.get("reforecast_bookings", 0.0)).style = "currency_style"
        ws.cell(row=row, column=10, value=summary.get("reforecast_gap_pct", 0.0)).style = "pct_style"
        row += 1

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 10)


def _build_pipeline_cohorts(wb: "Workbook", result: TieoutResult):
    """Build Pipeline Cohorts sheet — monthly pipeline creation & expected bookings."""
    ws = wb.create_sheet("Pipeline Cohorts")
    ws.sheet_properties.tabColor = "A371F7"  # purple

    scenario = _primary_scenario(result)

    ws["A1"] = "FY26 Pipeline Cohorts — Monthly Pipeline Created and Trajectory ARR"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    row = 4
    headers = [
        "Month", "Pipeline Created", "Trajectory Sales-Led ARR",
        "AE Capacity", "Capacity-Capped ARR", "Overflow",
        "Confidence", "Quarter"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    row += 1

    monthly_cap = scenario.monthly_capacity or []
    monthly_creation = list(getattr(scenario, "monthly_pipeline_creation", []) or [])
    expected_all = list(getattr(scenario, "monthly_bookings_expected", []) or [])
    capped_all = list(getattr(scenario, "monthly_bookings_capped", []) or [])
    overflow_all = list(getattr(scenario, "monthly_overflow", []) or [])

    # Confidence tier mapping: determine current quarter
    from datetime import date as _date
    # FY26 quarter boundaries
    QUARTER_DATES = {
        "Q1FY26": (_date(2026, 2, 1), _date(2026, 4, 30)),
        "Q2FY26": (_date(2026, 5, 1), _date(2026, 7, 31)),
        "Q3FY26": (_date(2026, 8, 1), _date(2026, 10, 31)),
        "Q4FY26": (_date(2026, 11, 1), _date(2027, 1, 31)),
    }

    def _get_quarter(d: _date) -> str:
        for q, (qs, qe) in QUARTER_DATES.items():
            if qs <= d <= qe:
                return q
        return "—"

    def _get_confidence(d: _date, quarters_from_now: int) -> str:
        # Use confidence_tier from quarterly data if available
        for q in scenario.quarters:
            qs, qe = QUARTER_DATES.get(q.quarter, (None, None))
            if qs and qe and qs <= d <= qe:
                return q.confidence_tier
        return "planned"
    monthly_capacity_vals = [getattr(mc, "ae_capacity", 0) or 0 for mc in monthly_cap]

    # Write one row per FY26 month
    for i, mc in enumerate(monthly_cap):
        month_date = getattr(mc, "month", None)
        month_label = getattr(mc, "label", str(month_date) if month_date else f"Month {i+1}")
        cap = getattr(mc, "ae_capacity", 0) or 0

        expected = expected_all[i] if i < len(expected_all) else 0
        capped = capped_all[i] if i < len(capped_all) else 0
        overflow = overflow_all[i] if i < len(overflow_all) else max(0, expected - capped)
        created = monthly_creation[i] if i < len(monthly_creation) else 0

        quarter_label = _get_quarter(month_date) if month_date else "—"
        tier = _get_confidence(month_date, i) if month_date else "planned"
        tier_color = _STATUS_COLOR.get(tier, CELL_FONT_COLOR)

        ws.cell(row=row, column=1, value=month_label).style = "label_style"
        ws.cell(row=row, column=2, value=created).style = "currency_style"
        ws.cell(row=row, column=3, value=expected).style = "currency_style"
        ws.cell(row=row, column=4, value=cap).style = "currency_style"
        ws.cell(row=row, column=5, value=capped).style = "currency_style"
        overflow_cell = ws.cell(row=row, column=6, value=overflow)
        overflow_cell.number_format = '$#,##0'
        overflow_cell.font = Font(color=YELLOW if overflow > 0 else CELL_FONT_COLOR)
        overflow_cell.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
        tier_cell = ws.cell(row=row, column=7, value=_display_confidence_tier(tier))
        tier_cell.font = Font(bold=True, color=tier_color)
        tier_cell.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
        ws.cell(row=row, column=8, value=quarter_label).style = "label_style"
        row += 1

    # Totals
    if monthly_cap:
        row += 1
        ws.cell(row=row, column=1, value="FY26 TOTAL").style = "header_style"
        ws.cell(row=row, column=2, value=sum(monthly_creation[:len(monthly_cap)])).style = "currency_style"
        ws.cell(row=row, column=3, value=sum(expected_all[:len(monthly_cap)])).style = "currency_style"
        ws.cell(row=row, column=4, value=sum(monthly_capacity_vals)).style = "currency_style"
        ws.cell(row=row, column=5, value=sum(capped_all[:len(monthly_cap)])).style = "currency_style"

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 8)


def _build_source_detail(wb: "Workbook", result: TieoutResult):
    """Build Source Detail sheet — source-stream drilldown by quarter and month."""
    ws = wb.create_sheet("Source Detail")
    ws.sheet_properties.tabColor = ACCENT_BLUE

    scenario = _primary_scenario(result)

    ws["A1"] = "FY26 Source Detail — Stream-Level Driver Drilldown"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:K1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    row = 4
    ws.cell(row=row, column=1, value="QUARTERLY STREAM SUMMARY").style = "subtitle_style"
    ws.merge_cells(f"A{row}:K{row}")
    row += 1

    summary_headers = [
        "Quarter", "Source Stream", "Input Metric", "Input / wk", "Expected S0 / wk",
        "Expected S1 / wk", "Expected S2 / wk", "Trajectory Quarter Pipeline",
        "Attributed Pipeline (warehouse)", "Attributed Opps (warehouse)", "Mode",
    ]
    for col, header in enumerate(summary_headers, 1):
        ws.cell(row=row, column=col, value=header).style = "header_style"
    row += 1

    for quarter in scenario.quarters:
        streams = (quarter.source_breakdown or {}).get("streams", {}) or {}
        mode = (quarter.source_breakdown or {}).get("mode", "config")
        for stream_data in streams.values():
            ws.cell(row=row, column=1, value=quarter.quarter).style = "label_style"
            ws.cell(row=row, column=2, value=_display_source_name(stream_data.get("display_name", ""))).style = "label_style"
            ws.cell(row=row, column=3, value=stream_data.get("input_label", "")).style = "label_style"
            ws.cell(row=row, column=4, value=stream_data.get("weekly_input", 0)).style = "num_style"
            ws.cell(row=row, column=5, value=stream_data.get("weekly_s0_count", 0)).style = "decimal_style"
            ws.cell(row=row, column=6, value=stream_data.get("weekly_s1_count", 0)).style = "decimal_style"
            ws.cell(row=row, column=7, value=stream_data.get("weekly_s2_count", 0)).style = "decimal_style"
            ws.cell(row=row, column=8, value=stream_data.get("quarter_pipeline_created", 0)).style = "currency_style"
            ws.cell(row=row, column=9, value=stream_data.get("actual_pipeline", 0)).style = "currency_style"
            ws.cell(row=row, column=10, value=stream_data.get("actual_opp_count", 0) or "").style = "num_style"
            ws.cell(row=row, column=11, value=_display_mode(mode)).style = "label_style"
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="MONTHLY STREAM DETAIL").style = "subtitle_style"
    ws.merge_cells(f"A{row}:K{row}")
    row += 1

    detail_headers = [
        "Month", "Quarter", "Source Stream", "Input Metric", "Input Volume",
        "Expected S0", "Expected S1", "Expected S2", "Pipeline Created", "Confidence", "Mode",
    ]
    for col, header in enumerate(detail_headers, 1):
        ws.cell(row=row, column=col, value=header).style = "header_style"
    row += 1

    for entry in getattr(scenario, "monthly_source_detail", []) or []:
        ws.cell(row=row, column=1, value=entry.get("month_label", "")).style = "label_style"
        ws.cell(row=row, column=2, value=entry.get("quarter", "")).style = "label_style"
        ws.cell(row=row, column=3, value=_display_source_name(entry.get("source", ""))).style = "label_style"
        ws.cell(row=row, column=4, value=entry.get("input_label", "")).style = "label_style"
        ws.cell(row=row, column=5, value=entry.get("input_count", 0)).style = "num_style"
        ws.cell(row=row, column=6, value=entry.get("s0_count", 0)).style = "decimal_style"
        ws.cell(row=row, column=7, value=entry.get("s1_count", 0)).style = "decimal_style"
        ws.cell(row=row, column=8, value=entry.get("s2_count", 0)).style = "decimal_style"
        ws.cell(row=row, column=9, value=entry.get("pipeline_created", 0)).style = "currency_style"
        ws.cell(row=row, column=10, value=_display_confidence_tier(entry.get("confidence_tier", ""))).style = "label_style"
        ws.cell(row=row, column=11, value=_display_mode(entry.get("mode", ""))).style = "label_style"
        row += 1

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 11)


def _build_expansion_detail(wb: "Workbook", result: TieoutResult):
    """Build Expansion Detail sheet — quarter-level expansion workstream."""
    ws = wb.create_sheet("Expansion Detail")
    ws.sheet_properties.tabColor = YELLOW

    scenario = _primary_scenario(result)

    ws["A1"] = "FY26 Expansion Detail — Standalone Existing-Customer Workstream"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:K1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    row = 4
    headers = [
        "Quarter",
        "Plan Expansion ARR",
        "Trajectory Expansion ARR",
        "Opening ARR Base",
        "Sales-Led Base ARR",
        "PLG Base ARR",
        "Renewal Expansion ARR",
        "Usage Expansion ARR",
        "PLG Expansion ARR",
        "Consumption True-Forward ARR",
        "Program Maturity",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).style = "header_style"
    row += 1

    for quarter in scenario.quarters:
        detail = getattr(quarter, "expansion_breakdown", None) or {}
        ws.cell(row=row, column=1, value=quarter.quarter).style = "label_style"
        ws.cell(row=row, column=2, value=quarter.td_expansion).style = "currency_style"
        ws.cell(row=row, column=3, value=quarter.bu_expansion_arr).style = "currency_style"
        ws.cell(row=row, column=4, value=detail.get("opening_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=5, value=detail.get("sales_led_base_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=6, value=detail.get("plg_base_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=7, value=detail.get("renewal_expansion_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=8, value=detail.get("usage_expansion_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=9, value=detail.get("plg_expansion_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=10, value=detail.get("consumption_true_forward_arr", 0.0)).style = "currency_style"
        ws.cell(row=row, column=11, value=detail.get("program_maturity_factor", 0.0)).style = "pct_style"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="FY26 TOTAL").style = "header_style"
    ws.cell(row=row, column=2, value=sum(q.td_expansion for q in scenario.quarters)).style = "currency_style"
    ws.cell(row=row, column=3, value=sum(q.bu_expansion_arr for q in scenario.quarters)).style = "currency_style"
    ws.cell(row=row, column=7, value=sum((q.expansion_breakdown or {}).get("renewal_expansion_arr", 0.0) for q in scenario.quarters)).style = "currency_style"
    ws.cell(row=row, column=8, value=sum((q.expansion_breakdown or {}).get("usage_expansion_arr", 0.0) for q in scenario.quarters)).style = "currency_style"
    ws.cell(row=row, column=9, value=sum((q.expansion_breakdown or {}).get("plg_expansion_arr", 0.0) for q in scenario.quarters)).style = "currency_style"
    ws.cell(row=row, column=10, value=sum((q.expansion_breakdown or {}).get("consumption_true_forward_arr", 0.0) for q in scenario.quarters)).style = "currency_style"

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 11)


def _build_monthly_capacity(wb: "Workbook", result: TieoutResult):
    """Build the Monthly Capacity sheet."""
    ws = wb.create_sheet("Monthly Capacity")
    ws.sheet_properties.tabColor = GREEN

    scenario = _primary_scenario(result)
    se_view_model = build_se_capacity_view_model(result)

    ws["A1"] = "FY26 Monthly Capacity Timeline"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:H1")

    row = 3
    headers = ["Month", "AEs Total", "AEs Ramped", "AEs Ramping",
               "Ramp %", "Monthly Capacity", "Monthly Target", "Capacity vs Target"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for mc in scenario.monthly_capacity:
        row += 1
        monthly_target = mc.monthly_target
        ws.cell(row=row, column=1, value=mc.label).style = "label_style"
        ws.cell(row=row, column=2, value=mc.ae_total).style = "num_style"
        ws.cell(row=row, column=3, value=mc.ae_ramped).style = "num_style"
        ws.cell(row=row, column=4, value=mc.ae_ramping).style = "num_style"
        ws.cell(row=row, column=5, value=mc.blended_ramp_pct).style = "pct_style"
        ws.cell(row=row, column=6, value=mc.ae_capacity).style = "currency_style"
        ws.cell(row=row, column=7, value=monthly_target if monthly_target is not None else "—").style = (
            "currency_style" if monthly_target is not None else "label_style"
        )

        if monthly_target is None:
            ws.cell(row=row, column=8, value="—").style = "label_style"
        else:
            ratio = mc.ae_capacity / monthly_target if monthly_target > 0 else 0
            style = "green_currency" if ratio >= 1 else "red_currency"
            ws.cell(row=row, column=8, value=mc.ae_capacity - monthly_target).style = style

    # Totals
    row += 1
    monthly_target_values = [
        mc.monthly_target for mc in scenario.monthly_capacity if mc.monthly_target is not None
    ]
    monthly_target_total = sum(monthly_target_values) if monthly_target_values else None
    ws.cell(row=row, column=1, value="FY26 TOTAL").style = "header_style"
    ws.cell(row=row, column=6, value=sum(
        mc.ae_capacity for mc in scenario.monthly_capacity
    )).style = "currency_style"
    ws.cell(row=row, column=7, value=monthly_target_total if monthly_target_total is not None else "—").style = (
        "currency_style" if monthly_target_total is not None else "label_style"
    )

    # ── SE Capacity Section ─────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="SE CAPACITY").style = "subtitle_style"
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    se_headers = ["Month", "SE Count (Roster)", "AE:SE Ratio", "Deals per SE"]
    for col, h in enumerate(se_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for se_row in se_view_model["monthly"]:
        row += 1
        ws.cell(row=row, column=1, value=se_row["label"]).style = "label_style"
        ws.cell(row=row, column=2, value=se_row["se_count"]).style = "num_style"
        ws.cell(row=row, column=3, value=se_row["ae_se_ratio"]).style = "label_style"
        ws.cell(row=row, column=4, value=se_row["deals_per_se"]).style = "decimal_style"

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 8)


def _build_scenario_comparison(wb: "Workbook", result: TieoutResult):
    """Build scenario comparison sheet if multiple scenarios exist."""
    if not result.scenarios:
        return
    archived = _archived_plan(result)
    comparison_contract = _comparison_contract()

    ws = wb.create_sheet("Scenarios")
    ws.sheet_properties.tabColor = "A371F7"

    ws["A1"] = "Scenario Comparison — Sales-Led Comparable View"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:F1")

    ws["A2"] = comparison_contract["note"]
    ws["A2"].style = "label_style"
    ws.merge_cells("A2:F2")

    row = 4
    headers = [
        "Scenario",
        "Sales-Led ARR",
        "Executive-Context Total Net New ARR",
        "FY Sales-Led Deficit vs Plan",
        "Sales-Led Deficit %",
        "Description",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    # Base scenario
    row += 1
    archived_sales_led_deficit = archived.fy_bookings_td - archived.fy_bookings_bu
    archived_sales_led_deficit_pct = archived_sales_led_deficit / archived.fy_bookings_td if archived.fy_bookings_td else 0
    ws.cell(row=row, column=1, value="Archived Plan").style = "label_style"
    ws.cell(row=row, column=2, value=archived.fy_bookings_bu).style = "currency_style"
    ws.cell(row=row, column=3, value=archived.fy_total_bu).style = "currency_style"
    ws.cell(row=row, column=4, value=archived_sales_led_deficit).style = (
        "red_currency" if archived_sales_led_deficit > 0 else "green_currency"
    )
    ws.cell(row=row, column=5, value=archived_sales_led_deficit_pct).style = "pct_style"
    ws.cell(row=row, column=6, value="Archived compatibility scenario").style = "label_style"

    # Other scenarios
    for name, scenario in result.scenarios.items():
        row += 1
        sales_led_deficit = scenario.fy_bookings_td - scenario.fy_bookings_bu
        sales_led_deficit_pct = sales_led_deficit / scenario.fy_bookings_td if scenario.fy_bookings_td else 0
        ws.cell(row=row, column=1, value=name).style = "label_style"
        ws.cell(row=row, column=2, value=scenario.fy_bookings_bu).style = "currency_style"
        ws.cell(row=row, column=3, value=scenario.fy_total_bu).style = "currency_style"
        ws.cell(row=row, column=4, value=sales_led_deficit).style = (
            "red_currency" if sales_led_deficit > 0 else "green_currency"
        )
        ws.cell(row=row, column=5, value=sales_led_deficit_pct).style = "pct_style"
        ws.cell(row=row, column=6, value=scenario.description).style = "label_style"

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 6)


def _build_scenario_planner_sheet(
    wb: "Workbook",
    result: TieoutResult,
    flexed_scenario: ScenarioResult = None,
    scenario_overrides: dict = None,
):
    """Build Scenario Planner sheet — only included when a flexed scenario is provided."""
    if flexed_scenario is None:
        return

    ws = wb.create_sheet("Scenario Planner")
    ws.sheet_properties.tabColor = "D2A8FF"  # light purple

    view_model = build_scenario_overlay_view_model(result, flexed_scenario=flexed_scenario)
    baseline = view_model["baseline"]

    ws["A1"] = f"Scenario Planner — {flexed_scenario.name}"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    if flexed_scenario.description:
        ws["A3"] = flexed_scenario.description
        ws["A3"].style = "label_style"
        ws.merge_cells("A3:H3")

    # ── Override Grid ───────────────────────────────────────────────
    row = 5
    ws.cell(row=row, column=1, value="APPLIED OVERRIDES").style = "subtitle_style"
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    override_rows = build_scenario_override_rows(scenario_overrides)
    if override_rows:
        override_headers = ["Override Parameter"] + list((scenario_overrides or {}).keys())
        for col, h in enumerate(override_headers, 1):
            ws.cell(row=row, column=col, value=h).style = "header_style"
        row += 1

        for override_row in override_rows:
            ws.cell(row=row, column=1, value=_display_label(override_row["parameter"])).style = "label_style"
            for col_idx, quarter_label in enumerate((scenario_overrides or {}).keys(), 2):
                val = override_row.get(quarter_label, "")
                if isinstance(val, float) and val < 1:
                    ws.cell(row=row, column=col_idx, value=val).style = "pct_style"
                elif isinstance(val, (int, float)):
                    ws.cell(row=row, column=col_idx, value=val).style = "num_style"
                else:
                    ws.cell(row=row, column=col_idx, value=str(val) if val != "" else "—").style = "label_style"
            row += 1
    else:
        if scenario_overrides:
            override_headers = ["Override Parameter"] + list(scenario_overrides.keys())
            for col, h in enumerate(override_headers, 1):
                ws.cell(row=row, column=col, value=h).style = "header_style"
            row += 1
        ws.cell(row=row, column=1, value="No overrides applied").style = "label_style"
        row += 1

    # ── Comparison Table ────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="SCENARIO COMPARISON").style = "subtitle_style"
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    comp_headers = ["Quarter", "Sales-Led Target", "Baseline", "Scenario", "Delta", "Delta %", "Gap to Sales-Led Target"]
    for col, h in enumerate(comp_headers, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"

    for row_data in view_model["quarterly"]:
        row += 1
        baseline_arr = row_data["baseline"]
        scenario_arr = row_data["scenario"]
        target = row_data["sales_led_target"]
        delta = row_data["delta"]
        delta_pct = row_data["delta_pct"] or 0
        gap_to_target = row_data["gap_to_target"]

        ws.cell(row=row, column=1, value=row_data["quarter"]).style = "label_style"
        ws.cell(row=row, column=2, value=target).style = "currency_style"
        ws.cell(row=row, column=3, value=baseline_arr).style = "currency_style"
        ws.cell(row=row, column=4, value=scenario_arr).style = "currency_style"
        delta_cell = ws.cell(row=row, column=5, value=delta)
        delta_cell.style = "green_currency" if delta >= 0 else "red_currency"
        ws.cell(row=row, column=6, value=delta_pct).style = "pct_style"
        ws.cell(row=row, column=7, value=gap_to_target).style = "red_currency" if gap_to_target < 0 else "green_currency"

    # FY total row
    row += 1
    fy_summary = view_model["fy_summary"]
    ws.cell(row=row, column=1, value="FY26 TOTAL").style = "header_style"
    ws.cell(row=row, column=2, value=fy_summary["target"]).style = "currency_style"
    ws.cell(row=row, column=3, value=fy_summary["baseline"]).style = "currency_style"
    ws.cell(row=row, column=4, value=fy_summary["scenario"]).style = "currency_style"
    fy_delta = fy_summary["scenario"] - fy_summary["baseline"]
    fy_delta_pct = fy_delta / fy_summary["baseline"] if fy_summary["baseline"] else 0
    fy_gap = fy_summary["scenario"] - fy_summary["target"]
    ws.cell(row=row, column=5, value=fy_delta).style = "green_currency" if fy_delta >= 0 else "red_currency"
    ws.cell(row=row, column=6, value=fy_delta_pct).style = "pct_style"
    ws.cell(row=row, column=7, value=fy_gap).style = "red_currency" if fy_gap < 0 else "green_currency"

    # ── Embedded Scenario Chart ──────────────────────────────────────
    if HAS_KALEIDO:
        row += 3
        ws.cell(row=row, column=1, value="SCENARIO CHART").style = "subtitle_style"
        ws.merge_cells(f"A{row}:H{row}")
        row += 1

        overlay_fig = _build_scenario_overlay_chart(baseline, flexed_scenario)
        overlay_stream = _render_chart_to_image(overlay_fig)
        if overlay_stream:
            img = XlImage(overlay_stream)
            img.width = 700
            img.height = 350
            ws.add_image(img, f"A{row}")
            row += 20

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 7)


def _build_data_health(wb: "Workbook", result: TieoutResult):
    """Build the Data Health sheet — freshness, bookings recon, decay, targets."""
    ws = wb.create_sheet("Data Health")
    ws.sheet_properties.tabColor = RED

    health = result.health_status or {}

    ws["A1"] = "Data Health — warehouse Audit & Reconciliation"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].style = "label_style"

    row = 4

    # ── Overall Status ──────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="OVERALL STATUS").style = "subtitle_style"
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    overall = health.get("overall_status", "unknown")
    ws.cell(row=row, column=1, value="Overall").style = "label_style"
    _status_label_cell(ws, row, 2, overall)
    row += 2

    # ── Freshness ───────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="FRESHNESS").style = "subtitle_style"
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    freshness_hdr = ["Mart / Source", "Last Updated", "Age (hours)", "Status"]
    for col, h in enumerate(freshness_hdr, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"
    row += 1

    freshness = health.get("freshness", {})
    if freshness:
        hours_old = freshness.get("hours_old")
        status = freshness.get("status", "unknown")
        msg = freshness.get("message", "")
        ws.cell(row=row, column=1, value="warehouse Marts").style = "label_style"
        ws.cell(row=row, column=2, value="(see warehouse)").style = "label_style"
        ws.cell(row=row, column=3, value=round(hours_old, 1) if hours_old is not None else "N/A").style = "num_style"
        _status_label_cell(ws, row, 4, status)
        row += 1
        if msg:
            ws.cell(row=row, column=1, value=msg[:120]).style = "label_style"
            ws.merge_cells(f"A{row}:E{row}")
            row += 1
    else:
        ws.cell(row=row, column=1, value="Warehouse unavailable — running in config-only mode").style = "label_style"
        ws.merge_cells(f"A{row}:E{row}")
        _status_label_cell(ws, row, 4, "yellow")
        row += 1

    row += 1

    # ── Bookings Reconciliation ─────────────────────────────────────────
    ws.cell(row=row, column=1, value="BOOKINGS RECONCILIATION").style = "subtitle_style"
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    bk_hdr = ["Source", "Amount", "Delta", "Delta %", "Status"]
    for col, h in enumerate(bk_hdr, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"
    row += 1

    bk = health.get("bookings_reconciliation", {})
    if bk and isinstance(bk, dict):
        cdw_val = bk.get("cdw_bookings")
        sf_val = bk.get("sf_bookings")
        delta_pct = bk.get("delta_pct", 0)
        status = bk.get("status", "ok")
        msg = bk.get("message", "")

        if cdw_val is not None:
            ws.cell(row=row, column=1, value="warehouse").style = "label_style"
            ws.cell(row=row, column=2, value=cdw_val).style = "currency_style"
            row += 1
        if sf_val is not None:
            ws.cell(row=row, column=1, value="Salesforce").style = "label_style"
            ws.cell(row=row, column=2, value=sf_val).style = "currency_style"
            delta = (cdw_val or 0) - (sf_val or 0)
            gap_style = "red_currency" if abs(delta) > 0.01 * max(abs(cdw_val or 1), 1) else "green_currency"
            ws.cell(row=row, column=3, value=delta).style = gap_style
            ws.cell(row=row, column=4, value=delta_pct).style = "pct_style"
            _status_label_cell(ws, row, 5, status)
            row += 1

        if msg:
            ws.cell(row=row, column=1, value=msg[:120]).style = "label_style"
            ws.merge_cells(f"A{row}:E{row}")
            row += 1
    else:
        ws.cell(row=row, column=1, value=bk.get("message", "Skipped")).style = "label_style"
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

    row += 1

    # ── Close Timing Assumption Check ──────────────────────────────────
    ws.cell(row=row, column=1, value="CLOSE TIMING ASSUMPTION CHECK").style = "subtitle_style"
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    dc_hdr = ["Metric", "Value", "", "Status", ""]
    for col, h in enumerate(dc_hdr, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"
    row += 1

    dc = health.get("decay_curve", {})
    if dc and isinstance(dc, dict):
        r2 = dc.get("r_squared")
        status = dc.get("status", "green")
        msg = dc.get("message", "")
        assumed = dc.get("assumed", [])
        actual = dc.get("actual", [])

        ws.cell(row=row, column=1, value="R² (goodness of fit)").style = "label_style"
        if r2 is not None:
            r2_cell = ws.cell(row=row, column=2, value=round(r2, 4))
            r2_cell.number_format = '0.0000'
            r2_cell.font = Font(color=CELL_FONT_COLOR)
            r2_cell.fill = PatternFill(start_color=CELL_BG, end_color=CELL_BG, fill_type="solid")
        _status_label_cell(ws, row, 4, status)
        row += 1

        if msg:
            ws.cell(row=row, column=1, value=msg[:120]).style = "label_style"
            ws.merge_cells(f"A{row}:E{row}")
            row += 1

        # Assumed vs actual distribution
        if assumed and actual:
            row += 1
            ws.cell(row=row, column=1, value="Assumed Curve").style = "label_style"
            ws.cell(row=row, column=2, value=str([round(x, 3) for x in assumed[:6]]) + "...").style = "label_style"
            row += 1
            ws.cell(row=row, column=1, value="Actual Curve").style = "label_style"
            ws.cell(row=row, column=2, value=str([round(x, 3) for x in actual[:6]]) + "...").style = "label_style"
            row += 1
    else:
        ws.cell(row=row, column=1, value=dc.get("message", "Using assumed curve (no actuals)")).style = "label_style"
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

    row += 1

    # ── Target Reconciliation ───────────────────────────────────────────
    ws.cell(row=row, column=1, value="TARGET RECONCILIATION (Q1FY26)").style = "subtitle_style"
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    tr_hdr = ["Source", "Value", "Delta", "Delta %", "Status"]
    for col, h in enumerate(tr_hdr, 1):
        ws.cell(row=row, column=col, value=h).style = "header_style"
    row += 1

    tr = health.get("targets", {})
    if tr and isinstance(tr, dict):
        status = tr.get("status", "aligned")
        delta_pct = tr.get("delta_pct", 0)
        msg = tr.get("message", "")

        ws.cell(row=row, column=1, value="YAML vs warehouse weekly sum").style = "label_style"
        ws.cell(row=row, column=4, value=delta_pct).style = "pct_style"
        _status_label_cell(ws, row, 5, status)
        row += 1

        if msg:
            ws.cell(row=row, column=1, value=msg[:120]).style = "label_style"
            ws.merge_cells(f"A{row}:E{row}")
            row += 1
    else:
        ws.cell(row=row, column=1, value=tr.get("message", "Skipped (Warehouse unavailable)") if isinstance(tr, dict) else "Skipped").style = "label_style"
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 5)


def _build_reviewer_notes(wb: "Workbook", result: TieoutResult):
    """Build Reviewer Notes sheet with caveats and run metadata."""
    ws = wb.create_sheet("Reviewer Notes")
    ws.sheet_properties.tabColor = YELLOW

    plan = result.top_down_plan or {}
    metadata = _build_review_metadata(result)
    scenario = _primary_scenario(result)
    comparison_contract = metadata.get("comparison_contract") or _comparison_contract()

    ws["A1"] = "Reviewer Notes — Caveats, Context, and Run Metadata"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:D1")

    row = 3
    ws.cell(row=row, column=1, value="RUN METADATA").style = "subtitle_style"
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    metadata_rows = [
        ("Generated At", metadata.get("generated_at", "")),
        ("Git SHA", metadata.get("git_sha", "")),
        ("Overall Health", metadata.get("overall_health", "unknown")),
        ("Top-Down Plan", plan.get("label", "")),
    ]
    for label, value in metadata_rows:
        ws.cell(row=row, column=1, value=label).style = "label_style"
        ws.cell(row=row, column=2, value=value).style = "label_style"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="REVIEWER NOTES").style = "subtitle_style"
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    notes = []
    if plan:
        notes.append(
            f"Top-down targets use '{plan.get('label', 'the configured plan')}' as a baseline reference rather than a claim that the current operating plan is unchanged."
        )
    notes.append(
        f"{comparison_contract['note']} Workbook comparisons should anchor on sales-led ARR for operator review."
    )

    derived_quarters = [q.quarter for q in scenario.quarters if q.is_derived_targets]
    if derived_quarters:
        notes.append(
            f"Operating targets for {', '.join(derived_quarters)} are derived from pipeline coverage, source mix, and recent stage relationships rather than a finance-approved later-quarter operating plan."
        )

    working_draft_quarters = [
        q.quarter
        for q in scenario.quarters
        if (getattr(q, "target_provenance", None) or {}).get("status") == "working_draft_explicit"
    ]
    if working_draft_quarters:
        notes.append(
            f"Operating targets for {', '.join(working_draft_quarters)} are explicit management-draft targets in config, but they are still not finance-approved."
        )

    provisional_recut_quarters = [
        q.quarter
        for q in scenario.quarters
        if (getattr(q, "target_provenance", None) or {}).get("status") == "provisional_recut_placeholder"
    ]
    if provisional_recut_quarters:
        notes.append(
            f"Operating targets for {', '.join(provisional_recut_quarters)} are provisional recut placeholders scaled from the prior operating plan. They remain directional planning references until finance-reviewed funnel targets are approved."
        )

    notes.append(
        "Expansion now runs through a standalone cohort-based workstream, but it is still assumption-driven and not yet calibrated from live renewal dates, consumption pricing rules, or warehouse expansion actuals."
    )

    modes = metadata.get("monthly_source_modes", [])
    if modes == ["config"]:
        notes.append(
            "This run used config-only trajectory logic. Review the outputs as directional planning estimates, not live-warehouse measurements."
        )
    elif "config" in modes and "cdw" in modes:
        notes.append(
            "This run mixes live-warehouse actuals for started quarters with config-driven projections for later quarters."
        )
    else:
        notes.append(
            "This run used live-warehouse actuals where available and projects later periods from the modeled funnel."
        )

    notes.append(
        "Source Detail is modeled at the stream level (Marketing / SDR, AE Self-Gen, PLG). It is appropriate for planning review, but not yet a channel-by-channel marketing attribution report."
    )

    for note in notes:
        ws.cell(row=row, column=1, value="-").style = "label_style"
        ws.cell(row=row, column=2, value=note).style = "label_style"
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    _auto_width(ws, max_width=50)
    _apply_cell_bg(ws, row + 2, 4)


def _build_assumptions(wb: "Workbook", result: TieoutResult):
    """Build the Assumptions sheet for auditability."""
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = BORDER_COLOR

    ws["A1"] = "Model Assumptions"
    ws["A1"].style = "title_style"
    ws.merge_cells("A1:C1")

    row = 3

    def write_section(title: str, items: list[tuple[str, any]]):
        nonlocal row
        ws.cell(row=row, column=1, value=title).style = "subtitle_style"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        for label, value in items:
            ws.cell(row=row, column=1, value=label).style = "label_style"
            cell = ws.cell(row=row, column=2, value=value)
            if isinstance(value, float) and value < 1:
                cell.style = "pct_style"
            elif isinstance(value, (int, float)):
                cell.style = "num_style"
            else:
                cell.style = "label_style"
            row += 1
        row += 1

    plan = result.top_down_plan or {}
    if plan:
        write_section("Top-Down Plan", [
            ("Label", plan.get("label", "")),
            ("Status", plan.get("status", "")),
            ("Derived From", plan.get("derived_from", "")),
            ("Source Workbook", plan.get("source_workbook", "")),
            ("Source Context", plan.get("source_context", "")),
            ("Reconciled On", plan.get("reconciled_on", "")),
        ])

    # Stage conversions
    sc = result.assumptions_snapshot.get("stage_conversion", {})
    write_section("Stage Conversion Rates", [
        ("S2 → Won", sc.get("s2_to_won", 0.18)),
        ("S3 → Won", sc.get("s3_to_won", 0.42)),
        ("S4 → Won", sc.get("s4_to_won", 0.58)),
        ("S5 → Won", sc.get("s5_to_won", 0.61)),
    ])

    # Funnel
    funnel = result.assumptions_snapshot.get("funnel", {})
    write_section("Funnel Conversion", [
        ("Lead → MQL", funnel.get("lead_to_mql", 0.30)),
        ("MQL → S0", funnel.get("mql_to_s0", 0.15)),
        ("S0 → S1", funnel.get("s0_to_s1", 0.55)),
        ("S1 → S2", funnel.get("s1_to_s2", 0.25)),
    ])

    # Capacity
    cap = result.assumptions_snapshot.get("capacity", {})
    quotas = cap.get("quotas", {})
    write_section("Capacity", [
        ("Enterprise AE Quota", quotas.get("enterprise_ae_annual", 1_400_000)),
        ("Mid-Market AE Quota", quotas.get("midmarket_ae_annual", 650_000)),
        ("Attainment Rate", cap.get("attainment_rate", 0.80)),
        ("Pipeline Coverage", cap.get("pipeline_coverage_ratio", 3.5)),
    ])

    # ── Conversion Rate Methodology  ──────────────────────
    scenario = _primary_scenario(result)
    adr007_items = [
        ("Methodology", "Weighted blend of mature monthly cohorts "),
    ]
    # Extract blended S0→S1 and S1→S2 rates from first quarter's conversion_rates
    q1 = scenario.quarters[0] if scenario.quarters else None
    if q1:
        cr = q1.conversion_rates or {}
        for stage_key, label in [("s0_to_s1", "S0 → S1 Blended Rate"), ("s1_to_s2", "S1 → S2 Blended Rate")]:
            stage_data = cr.get(stage_key, {})
            if isinstance(stage_data, dict):
                # Could be nested by source or flat
                for source, sdata in stage_data.items():
                    if isinstance(sdata, dict) and "rate" in sdata:
                        rate = sdata.get("rate", 0)
                        n = sdata.get("n", 0)
                        src = _display_source_name(source)
                        adr007_items.append((f"{label} ({src})", rate))
                        adr007_items.append((f"{label} ({src}) Sample n", n))
                        break

    write_section("Conversion Rate Methodology ", adr007_items)

    _auto_width(ws)
    _apply_cell_bg(ws, row + 2, 3)


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_tieout_workbook(
    result: TieoutResult,
    output_path: Optional[Union[str, Path]] = None,
    flexed_scenario: ScenarioResult = None,
    scenario_overrides: dict = None,
) -> Union[Path, io.BytesIO]:
    """
    Export the tie-out result as a multi-sheet Excel workbook.

    Sheets (v3):
        1. Executive Summary    — KPIs, health banner, funnel gap summary
        2. Bookings Bridge      — Monthly existing vs future wins breakdown
        3. Funnel Tie-Out       — Stage-by-stage plan vs actual/projected + QTD pacing
        4. Operating Reforecast — Actuals vs plan pace vs BU
        5. Source Detail        — Monthly driver drilldown by source stream
        6. Expansion Detail     — Quarter-level expansion workstream detail
        7. Pipeline Cohorts     — Monthly cohort decay & bookings projection
        8. Monthly Capacity     — Headcount, capacity timeline, SE section
        9. Scenario Planner     — Optional: per-quarter overrides and comparison
        10. Scenarios           — Only if multiple scenarios computed
        11. Data Health         — warehouse freshness, reconciliation, decay fit
        12. Reviewer Notes      — Caveats, metadata, and review guidance
        13. Assumptions         — Full model parameters + ARCHITECTURE.md methodology

    Args:
        result: TieoutResult from PlanningTieout.compute_full()
        output_path: File path to save to. If None, returns BytesIO buffer.
        flexed_scenario: Optional ScenarioResult from the scenario planner.
        scenario_overrides: Optional dict of per-quarter overrides applied.

    Returns:
        Path to saved file, or BytesIO buffer if no path given.
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    wb = Workbook()
    _setup_styles(wb)

    _build_executive_summary(wb, result)
    _build_bookings_bridge(wb, result)
    _build_funnel_tieout(wb, result)
    _build_operating_reforecast(wb, result)
    _build_source_detail(wb, result)
    _build_expansion_detail(wb, result)
    _build_pipeline_cohorts(wb, result)
    _build_monthly_capacity(wb, result)
    _build_scenario_planner_sheet(wb, result, flexed_scenario, scenario_overrides)
    _build_scenario_comparison(wb, result)
    _build_data_health(wb, result)
    _build_reviewer_notes(wb, result)
    _build_assumptions(wb, result)

    if output_path:
        output_path = Path(output_path)
        wb.save(output_path)
        return output_path
    else:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


def export_tieout_json(
    result: TieoutResult,
    output_path: Optional[Union[str, Path]] = None,
    flexed_scenario: ScenarioResult = None,
    scenario_overrides: dict = None,
) -> Union[Path, io.BytesIO]:
    """Export the tie-out result as a JSON snapshot with run metadata."""
    scenario = _primary_scenario(result)
    bookings_bridge_vm = build_bookings_bridge_view_model(result)
    se_capacity_vm = build_se_capacity_view_model(result)
    scenario_vm = build_scenario_overlay_view_model(result, flexed_scenario=flexed_scenario)

    bookings_bridge_monthly = [
        {
            "month": row["month"].isoformat() if hasattr(row["month"], "isoformat") else str(row["month"]),
            "existing_wins": row["existing_wins"],
            "future_wins": row["future_wins"],
            "total": row["total_uncapped"],
            "target": row["target"],
            "gap": row["gap"],
        }
        for row in bookings_bridge_vm["monthly"]
    ]
    bookings_bridge_quarterly = [
        {
            "quarter": row["quarter"],
            "target": row["target"],
            "trajectory": row["trajectory"],
            "gap": row["gap"],
        }
        for row in bookings_bridge_vm["quarterly"]
    ]

    pacing = []
    for q in scenario.quarters:
        pacing_vm = build_funnel_pacing_view_model(q, as_of=date.today())
        if pacing_vm["quarter_state"] != "in_progress":
            continue
        for row in pacing_vm["rows"]:
            pacing.append(
                {
                    "quarter": q.quarter,
                    "stage": row["stage"],
                    "weekly_target": row["weekly_target"],
                    "weekly_actual": row["weekly_actual"],
                    "qtd_target": row["qtd_target"],
                    "qtd_actual": row["qtd_actual"],
                    "pacing_pct": row["pacing_pct"],
                }
            )

    payload = {
        "metadata": _build_review_metadata(result),
        "result": {
            "beginning_arr": result.beginning_arr,
            "beginning_arr_provenance": result.beginning_arr_provenance,
            "bookings_summary": result.bookings_summary,
            "bookings_summary_provenance": result.bookings_summary_provenance,
            "health_status": result.health_status,
            "top_down_plan": result.top_down_plan,
            "assumptions_snapshot": result.assumptions_snapshot,
            "trajectory": scenario.to_dict(),
            "archived_plan": _archived_plan(result).to_dict(),
            "base": result.base.to_dict(),
            "scenarios": {name: s.to_dict() for name, s in result.scenarios.items()},
        },
        "bookings_bridge": {
            "monthly": bookings_bridge_monthly,
            "quarterly": bookings_bridge_quarterly,
        },
        "se_capacity": {
            "roster_se_count": se_capacity_vm["current_ses"],
            "incoming_se_count": se_capacity_vm["incoming_ses"],
            "plan_se_target": se_capacity_vm["plan_se_target"],
            "active_s2_plus_deals": se_capacity_vm["active_s2_plus_deals"],
            "deals_per_se": se_capacity_vm["deals_per_se"],
            "monthly": se_capacity_vm["monthly"],
        },
        "pacing": pacing,
    }

    # Include scenario planner if present
    if flexed_scenario:
        payload["scenario"] = {
            "name": flexed_scenario.name,
            "description": flexed_scenario.description,
            "overrides": scenario_overrides,
            "override_rows": build_scenario_override_rows(scenario_overrides),
            "quarterly": scenario_vm["quarterly"],
            "fy_summary": {
                "bookings_td": scenario_vm["fy_summary"]["target"],
                "bookings_bu": scenario_vm["fy_summary"]["scenario"],
                "gap": scenario_vm["fy_summary"]["gap"],
                "gap_pct": scenario_vm["fy_summary"]["gap_pct"],
            },
        }

    encoded = json.dumps(payload, indent=2, default=str).encode("utf-8")

    if output_path:
        output_path = Path(output_path)
        output_path.write_bytes(encoded)
        return output_path

    buffer = io.BytesIO(encoded)
    buffer.seek(0)
    return buffer
